import os

from openpyxl import load_workbook


class ExcelTool:

    def __init__(self, fileName):
        self._xlsx = None
        self.fileName = os.path.abspath(fileName)

    def getWorkbook(self):
        print(self.fileName)
        self._xlsx = load_workbook(filename=self.fileName, data_only=True)
        return True

    def getXLSX(self):
        return self._xlsx

    def close(self):
        if self._xlsx:
            self._xlsx.close()

    def getSheetNames(self):
        return self._xlsx.sheetnames

    def getSheetCount(self):
        return len(self._xlsx.sheetnames)

    def getSheetNameByIndex(self, index):
        return self._xlsx.sheetnames[index]

    def getSheetByIndex(self, index):
        return self._xlsx[self.getSheetNameByIndex(index)]

    def getSheetBySheetName(self, sheetName):
        return self._xlsx[sheetName]

    def getRowCount(self, sheetName):
        return self._xlsx[sheetName].max_row

    def getColCount(self, sheetName):
        return self._xlsx[sheetName].max_column

    def getValue(self, sheet, row, col):
        """
        row从0开始，col从0开始
        """
        return sheet.cell(row + 1, col + 1).value

    def getRowValues(self, sheet, row):
        iters = self._getSheetRowIters(sheet, row + 1)
        return [e.value for e in iters]

    def _getSheetRowIters(self, sheet, row):
        for row in sheet.iter_rows(min_row=row, max_row=row):
            for cell in row:
                yield cell

    def _getSheetColIters(self, sheet, col):
        for col in sheet.iter_rows(min_col=col, max_col=col):
            for cell in col:
                yield cell

    def getColValues(self, sheet, col):
        iters = self._getSheetColIters(sheet, col + 1)
        return [e.value for e in iters]


def main():
    tool = ExcelTool('../test/skills.xlsx')
    tool.getWorkbook()
    print(tool.getSheetCount())

    sheetCount = tool.getSheetCount()
    for i in range(sheetCount):
        print(tool.getSheetNameByIndex(i))

    sheet = tool.getSheetByIndex(0)
    values = tool.getRowValues(sheet, 1)
    print(values)
    values = tool.getColValues(sheet, 1)
    print(values)

    tool.close()


if __name__ == '__main__':
    main()
